"""
사업수지표 범용 추출기 v3
─ 병합 셀 기반 계층 인식
─ 텍스트 컬럼 위치 → 들여쓰기 레벨 (대/중/소분류)
─ 스택 알고리즘으로 트리 빌드
─ 어떤 양식이든 항목 그대로 추출
"""

import json, os, sys, re, zipfile, tempfile

# ────────────────────────────────────────────────────────
# 상수
# ────────────────────────────────────────────────────────
_SUBTOTAL_RE = re.compile(r'소\s*계|합\s*계$|소\s*합\s*계')
_TOTAL_KW = {
    '매출합계','매출계','총매출','수입합계','수입계',
    '지출합계','지출계','사업비합계','총사업비',
    '세전이익','당기순이익','예상손익','손익'
}
_SKIP_RE = re.compile(r'^\s*(no\.?|구\s*분|항\s*목|내\s*역|비\s*고|단\s*위|금\s*액|비\s*율|합\s*계금액|순번)\s*$', re.IGNORECASE)


def _n(s):
    """공백 제거 정규화."""
    return re.sub(r'\s+', '', str(s or '').strip())


# ────────────────────────────────────────────────────────
# 1. 워크시트 로드 (병합 셀 마스킹 포함)
# ────────────────────────────────────────────────────────
def _strip_custom_props(src):
    """docProps/custom.xml 제거 임시 xlsx 생성."""
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.close()
    with zipfile.ZipFile(src, 'r') as zin, \
         zipfile.ZipFile(tmp.name, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            if item.filename != 'docProps/custom.xml':
                zout.writestr(item, zin.read(item.filename))
    return tmp.name


_SHEET_PRIORITY = ['사업수지', '수지표', '사업수지표']

def _pick_sheet_xls(wb):
    names = wb.sheet_names()
    for pname in _SHEET_PRIORITY:
        for i, n in enumerate(names):
            if pname in n:
                return wb.sheet_by_index(i)
    return wb.sheet_by_index(0)

def _pick_sheet_xlsx(wb):
    for pname in _SHEET_PRIORITY:
        for n in wb.sheetnames:
            if pname in n:
                return wb[n]
    return wb.active


def load_sheet(file_path):
    """
    반환: (rows, merge_map)
      rows      : 0-based, 병합 계속셀은 None 처리
      merge_map : {(r,c): (r0,c0)} — 병합 범위의 시작 셀 위치
    """
    merge_map = {}

    if file_path.lower().endswith('.xlsx'):
        try:
            import openpyxl
        except ModuleNotFoundError as e:
            raise RuntimeError('openpyxl 라이브러리가 필요합니다.') from e

        tmp = None
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception:
            try:
                tmp = _strip_custom_props(file_path)
                wb = openpyxl.load_workbook(tmp, data_only=True)
            finally:
                if tmp and os.path.exists(tmp):
                    os.unlink(tmp)

        ws = _pick_sheet_xlsx(wb)
        for mr in ws.merged_cells.ranges:
            r0, c0 = mr.min_row - 1, mr.min_col - 1
            for r in range(mr.min_row - 1, mr.max_row):
                for c in range(mr.min_col - 1, mr.max_col):
                    if r != r0 or c != c0:
                        merge_map[(r, c)] = (r0, c0)

        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        wb.close()

    else:
        try:
            import xlrd
        except ModuleNotFoundError as e:
            raise RuntimeError('xls 파일 분석에는 xlrd 라이브러리가 필요합니다. xlsx로 저장 후 다시 업로드하세요.') from e

        wb = xlrd.open_workbook(file_path)
        ws = _pick_sheet_xls(wb)
        for rlo, rhi, clo, chi in ws.merged_cells:
            for r in range(rlo, rhi):
                for c in range(clo, chi):
                    if r != rlo or c != clo:
                        merge_map[(r, c)] = (rlo, clo)
        rows = []
        for ri in range(ws.nrows):
            row = []
            for ci in range(ws.ncols):
                if (ri, ci) in merge_map:
                    row.append(None)
                else:
                    v = ws.cell_value(ri, ci)
                    row.append(v if v != '' else None)
            rows.append(row)

    return rows, merge_map


# ────────────────────────────────────────────────────────
# 2. 금액 컬럼 & 단위 감지
# ────────────────────────────────────────────────────────
def detect_amount_info(rows):
    """반환: (amount_col 0-based or None, unit_multiplier)"""
    unit_mult = 1

    for row in rows[:15]:
        for cell in row:
            if cell and isinstance(cell, str):
                nc = _n(cell)
                if '백만원' in nc:
                    unit_mult = 1000   # 백만원 → ×1000 = 천원 단위로 통일
                    break
                elif '천원' in nc:
                    unit_mult = 1
                    break

    # 금액 헤더로 컬럼 탐색
    amt_exact = {'금액', '금액(천원)', '금(천원)', '금액천원', '금액(원)', '합계금액', '예산금액'}
    for r_idx, row in enumerate(rows[:30]):
        for c_idx, cell in enumerate(row):
            if cell and isinstance(cell, str):
                nc = _n(cell)
                if nc in amt_exact:
                    return c_idx, unit_mult

    # 폴백: 숫자가 가장 많이 등장하는 컬럼
    col_hits = {}
    for row in rows[5:min(len(rows), 60)]:
        for c_idx, cell in enumerate(row):
            if isinstance(cell, (int, float)) and abs(cell) >= 10000:
                col_hits[c_idx] = col_hits.get(c_idx, 0) + 1
    if col_hits:
        return max(col_hits, key=col_hits.get), unit_mult

    return None, unit_mult


# ────────────────────────────────────────────────────────
# 3. 산출 내역 컬럼 감지 & 수식 텍스트 추출
# ────────────────────────────────────────────────────────
def detect_formula_col(rows):
    """'산출 내역' 헤더 컬럼 인덱스 반환 (없으면 None)."""
    for row in rows[:15]:
        for c_idx, cell in enumerate(row):
            if cell and isinstance(cell, str):
                nc = _n(cell)
                if '산출' in nc and '내역' in nc:
                    return c_idx
    return None


_FORMULA_SKIP = {'산출내역', '비고', '비율', '세대가격:', '비  고', '비  율', '비고란'}
_EXCEL_ERR    = re.compile(r'^#')  # #DIV/0!, #REF!, #N/A 등

def _fmt_num(v):
    """산출 내역용 숫자 포맷."""
    av = abs(v)
    if av == int(av):
        return f'{int(av):,}'
    if av >= 100:
        return f'{av:,.1f}'
    return f'{av:.2f}'


def extract_row_formula_text(row, formula_col, merge_map, r_idx):
    """
    '산출 내역' 영역(formula_col ~ formula_col+16)에서
    사람이 읽기 쉬운 산출식 텍스트 추출.
    예: '33.75평 × 205세대 × 12,791천원'
    반환 None 조건:
      - × 가 없거나 (단순 레이블)
      - 첫 토큰이 숫자가 아닌 경우 (기준값 없이 비율만 있는 경우)
      - Excel 오류값(#DIV/0! 등) 포함
    """
    if formula_col is None or formula_col >= len(row):
        return None

    parts = []
    has_error = False

    for c_idx in range(formula_col, min(len(row), formula_col + 17)):
        if (r_idx, c_idx) in merge_map:
            continue
        cell = row[c_idx]
        if cell is None:
            continue

        if isinstance(cell, str):
            s = cell.strip()
            if not s:
                continue
            if _EXCEL_ERR.match(s):      # #DIV/0! 등 오류값 제거
                has_error = True
                continue
            if _n(s) in _FORMULA_SKIP:
                continue
            if s.lower() == 'x':
                parts.append('×')
            else:
                parts.append(s)

        elif isinstance(cell, (int, float)) and cell != 0:
            av = abs(cell)
            if av > 500_000 or av < 0.001:
                continue
            parts.append(_fmt_num(cell))

    if not parts:
        return None

    text = ' '.join(parts)
    text = re.sub(r'(×\s*)+', '×', text)
    text = re.sub(r'^\s*×\s*', '', text)
    text = re.sub(r'\s*×\s*$', '', text)
    text = text.strip()

    if not text or text == '0':
        return None
    if '×' not in text:           # 곱셈식이 없으면 산출식 아님
        return None
    if not text[0].isdigit():     # "천원 ×0.03" 처럼 기준값 없이 비율만 있는 경우 제거
        return None
    if has_error:                  # 오류값 포함된 행은 제거
        return None
    # 끝에 단위 레이블만 남은 경우 제거: "×천원", "×세대" 등
    text = re.sub(r'\s*×\s*(천원|만원|원|평|세대|개월|㎡|년)\s*$', '', text).strip()
    if '×' not in text:
        return None
    return text


# ────────────────────────────────────────────────────────
# 4. 행 단위 추출 (flat list)
# ────────────────────────────────────────────────────────
def _row_amount(row, amount_col, unit_mult):
    if amount_col is None or amount_col >= len(row):
        return 0
    v = row[amount_col]
    if isinstance(v, (int, float)) and v != 0:
        return round(v * unit_mult)
    return 0


def _row_texts(row, merge_map, r_idx):
    """병합 계속 셀을 제외한 (col_idx, text) 목록 반환."""
    result = []
    for c_idx, cell in enumerate(row[:20]):
        if cell is None:
            continue
        if (r_idx, c_idx) in merge_map:
            continue
        if not isinstance(cell, str):
            continue
        # 다중행 텍스트(\n) 정규화
        txt = re.sub(r'[\r\n]+', ' ', cell.strip())
        txt = re.sub(r'\s{2,}', ' ', txt).strip()
        if len(_n(txt)) < 2:
            continue
        if _SKIP_RE.match(txt):
            continue
        result.append((c_idx, txt))
    return result


def _detect_type(label):
    nc = _n(label)
    if nc in _TOTAL_KW or any(kw in nc for kw in _TOTAL_KW):
        return 'total'
    if _SUBTOTAL_RE.search(nc):
        return 'subtotal'
    return 'item'


def extract_flat(rows, merge_map, amount_col, unit_mult, formula_col=None):
    """
    각 행에서 (level, label, amount, type, formula) 추출.
    같은 행에 여러 텍스트가 있으면 각각 노드로 분리:
      좌측(낮은 컬럼) = 섹션 헤더 (amount=0)
      우측(높은 컬럼) = 항목 (amount=실제값, formula 배정)
    """
    flat = []

    for r_idx, row in enumerate(rows):
        texts = _row_texts(row, merge_map, r_idx)
        if not texts:
            continue

        amount  = _row_amount(row, amount_col, unit_mult)
        formula = extract_row_formula_text(row, formula_col, merge_map, r_idx)

        if len(texts) == 1:
            c_idx, label = texts[0]
            dtype = _detect_type(label)
            flat.append({
                'level':   c_idx,
                'label':   label,
                'amount':  amount,
                'type':    dtype,
                'formula': formula if dtype == 'item' else None,
                'r':       r_idx,
            })
        else:
            # 여러 텍스트: 맨 마지막(가장 오른쪽)에만 금액·산출식 배정
            for i, (c_idx, label) in enumerate(texts):
                is_last = (i == len(texts) - 1)
                dtype   = _detect_type(label)
                flat.append({
                    'level':   c_idx,
                    'label':   label,
                    'amount':  amount if is_last else 0,
                    'type':    dtype,
                    'formula': (formula if dtype == 'item' else None) if is_last else None,
                    'r':       r_idx,
                })

    return flat


# ────────────────────────────────────────────────────────
# 4. 트리 빌드
# ────────────────────────────────────────────────────────
class Node:
    __slots__ = ('label', 'amount', 'node_type', 'children', 'formula')

    def __init__(self, label, amount, node_type, formula=None):
        self.label = label
        self.amount = amount
        self.node_type = node_type
        self.children = []
        self.formula = formula

    def effective(self):
        """
        직접 금액이 있으면 사용.
        없으면: 소계/합계 자식이 있으면 마지막 소계 사용 (중복 합산 방지).
        없으면: 비소계 자식들만 합산.
        """
        if self.amount != 0:
            return abs(self.amount)
        # 소계/합계 자식 중 마지막 것이 권위 있는 합계
        sub_children = [c for c in self.children
                        if c.node_type in ('subtotal', 'total')]
        if sub_children:
            return sub_children[-1].effective()
        return sum(c.effective() for c in self.children)

    def to_dict(self):
        d = {
            'label':     self.label,
            'amount':    self.amount,
            'effective': round(self.effective()),
            'type':      self.node_type,
            'children':  [c.to_dict() for c in self.children],
        }
        if self.formula:
            d['formula'] = self.formula
        return d


def build_tree(flat):
    """스택 기반 트리 빌드."""
    roots = []
    stack = [(-1, None)]   # (level, Node|None)

    for row in flat:
        level = row['level']
        node  = Node(row['label'], row['amount'], row['type'], row.get('formula'))

        while len(stack) > 1 and stack[-1][0] >= level:
            stack.pop()

        parent = stack[-1][1]
        if parent is None:
            roots.append(node)
        else:
            parent.children.append(node)

        stack.append((level, node))

    return roots


_MIN_AMOUNT = 1000   # 1,000천원 = 100만원 미만은 비율/메타데이터로 간주
_META_RE = re.compile(r'수익율|수익률|평\s*당|㎡\s*당|단\s*가|이\s*율|비\s*율|수\s*익\s*율|수\s*익\s*률')

def prune_tree(nodes):
    """금액도 없고 자식도 없는 빈 노드, 비율값/단가 메타 노드 제거."""
    result = []
    for node in nodes:
        node.children = prune_tree(node.children)
        eff = node.effective()
        if not node.children:
            if eff < _MIN_AMOUNT:
                continue  # 100만원 미만 단독 노드 제거 (비율값, 면적 등)
            if _META_RE.search(_n(node.label)):
                continue  # 수익율/단가 메타 노드 제거
        elif eff <= 0:
            continue
        result.append(node)
    return result


# ────────────────────────────────────────────────────────
# 5. KPI values 추출 (대시보드 요약용)
# ────────────────────────────────────────────────────────
_SEC_KW = {
    'land':    ['토지비'],
    'constr':  ['건축비', '공사비'],
    'sales':   ['판매비', '분양비'],
    'levy':    ['부담금'],
    'incid':   ['부대비'],
    'finance': ['금융비', '금융비용', '이자'],
}
_REV_KW  = ['매출합계','수입합계','총매출','분양수입합계','매출계','수입계']
_COST_KW = ['사업비합계','지출합계','총사업비','총지출','지출계']


def extract_values(roots):
    v = {k: 0 for k in ['rev','cost','land','constr','sales','levy','incid','finance','profit','roi']}

    # 1차: 키워드로 값 탐색 (전체 트리 재귀)
    def walk(nodes):
        for nd in nodes:
            nc = _n(nd.label)
            amt = nd.effective()
            if amt <= 0:
                walk(nd.children)
                continue
            if any(kw in nc for kw in _REV_KW) and v['rev'] == 0:
                v['rev'] = amt
            if any(kw in nc for kw in _COST_KW) and v['cost'] == 0:
                v['cost'] = amt
            for key, kws in _SEC_KW.items():
                if v[key] == 0 and any(kw in nc for kw in kws):
                    v[key] = amt
                    break
            walk(nd.children)

    walk(roots)

    # 2차 폴백: "수입/매출" 섹션 헤더 → rev, "지출/사업비" → cost
    # (Type A 처럼 별도 합계 행 없이 섹션 자체가 합계인 경우)
    _REV_SECTION  = {'수입', '매출', '수  입', '매  출'}
    _COST_SECTION = {'지출', '사업비', '지  출', '사  업  비'}
    for nd in roots:
        nc = _n(nd.label)
        amt = nd.effective()
        if v['rev'] == 0 and nc in _REV_SECTION and amt > 0:
            v['rev'] = amt
        if v['cost'] == 0 and nc in _COST_SECTION and amt > 0:
            v['cost'] = amt

    # 3차: 섹션 소계가 없으면 지출 섹션 자식에서 추출
    if sum(v[k] for k in ['land','constr','sales','levy','incid','finance']) == 0:
        for nd in roots:
            nc = _n(nd.label)
            if nc in {'지출', '지  출', '사업비'} or any(kw in nc for kw in _COST_KW):
                for child in nd.children:
                    sec = _match_sec(child.label)
                    if sec and v[sec] == 0:
                        v[sec] = child.effective()

    if v['cost'] == 0:
        v['cost'] = sum(v[k] for k in ['land','constr','sales','levy','incid','finance'])
    if v['rev'] > 0 and v['cost'] > 0:
        v['profit'] = v['rev'] - v['cost']
        v['roi'] = round(v['profit'] / v['rev'] * 100, 2)
    return v


def _match_sec(label):
    nc = _n(label)
    for key, kws in _SEC_KW.items():
        if any(kw in nc for kw in kws):
            return key
    return None


def tree_to_flat_items(roots, depth=0):
    """트리 → flat line_items (하위 호환용)."""
    items = []
    for nd in roots:
        items.append({
            'label': nd.label,
            'amount': round(nd.effective()),
            'type': nd.node_type,
            'depth': depth,
            'has_children': bool(nd.children),
        })
        items.extend(tree_to_flat_items(nd.children, depth + 1))
    return items


# ────────────────────────────────────────────────────────
# 6. Excel 진입점
# ────────────────────────────────────────────────────────
def normalize_name(raw):
    s = re.sub(r'\.(pdf|xlsx|xls)$', '', raw, flags=re.IGNORECASE)
    s = re.sub(r'^\d+\.\s*', '', s)
    s = re.sub(r'^사업수지\s*\(?', '', s)
    s = re.sub(r'\)\s*$', '', s)
    return s.strip(' -_') or '신규 현장'


def extract_from_excel(file_path, original_name=None):
    try:
        rows, merge_map = load_sheet(file_path)
        name = normalize_name(original_name or os.path.basename(file_path))
        amount_col, unit_mult = detect_amount_info(rows)
        formula_col = detect_formula_col(rows)
        flat = extract_flat(rows, merge_map, amount_col, unit_mult, formula_col)
        roots = build_tree(flat)
        roots = prune_tree(roots)
        values = extract_values(roots)
        line_items = tree_to_flat_items(roots)
        tree_dicts = [nd.to_dict() for nd in roots]

        return {
            'success': True,
            'name': name,
            'data': {
                'name': name,
                'tree': tree_dicts,
                'line_items': line_items,
                'rev_details': [],
                'values': values,
            }
        }
    except Exception as e:
        import traceback
        return {'success': False, 'error': f'Excel 분석 실패: {e}\n{traceback.format_exc()}'}


# ────────────────────────────────────────────────────────
# 7. PDF 추출기
# ────────────────────────────────────────────────────────
_PDF_KW = {
    'rev':     ['매출합계','매출계','총매출','분양수입합계'],
    'cost':    ['지출합계','지출계','총사업비','사업비합계'],
    'land':    ['토지비소계','토지비계','토지비','용지비'],
    'constr':  ['건축비소계','건축비계','공사비소계','건축비','공사비'],
    'sales':   ['판매비소계','판매비계','판매비'],
    'finance': ['금융비용','금융비','이자비용'],
    'incid':   ['부대비용','기타경비','부대비'],
    'levy':    ['부담금소계','부담금'],
}


def _parse_nums(text):
    nums = []
    for m in re.finditer(r'-?\d{1,3}(?:,\d{3})+', text):
        try:
            nums.append(float(m.group().replace(',', '')))
        except ValueError:
            pass
    if nums:
        return nums
    for m in re.finditer(r'-?\d{5,}', text):
        try:
            nums.append(float(m.group()))
        except ValueError:
            pass
    return nums


def _extract_pdf_lines(file_path):
    try:
        import pdfplumber
    except ModuleNotFoundError:
        pdfplumber = None

    if pdfplumber:
        lines = []
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                words = page.extract_words()
                if not words:
                    continue
                line_map = {}
                for w in words:
                    y = round(w['top'])
                    line_map.setdefault(y, []).append(w)
                for y in sorted(line_map):
                    row = sorted(line_map[y], key=lambda x: x['x0'])
                    lines.append(' '.join(w['text'] for w in row))
        return lines

    try:
        from pypdf import PdfReader
    except ModuleNotFoundError as e:
        raise RuntimeError('PDF 분석에는 pdfplumber 또는 pypdf 라이브러리가 필요합니다.') from e

    lines = []
    reader = PdfReader(file_path)
    for page in reader.pages:
        text = page.extract_text() or ''
        for line in text.splitlines():
            line = line.strip()
            if line:
                lines.append(line)
    return lines


def extract_from_pdf(file_path, original_name=None):
    try:
        name = normalize_name(original_name or os.path.basename(file_path))
        lines = _extract_pdf_lines(file_path)

        values = {k: 0 for k in ['rev','cost','land','constr','sales','levy','incid','finance','profit','roi']}
        tree_items = []
        matched = set()
        for line in lines:
            nc = _n(line)
            for key, kws in _PDF_KW.items():
                if key in matched:
                    continue
                for kw in kws:
                    if kw in nc:
                        nums = _parse_nums(line)
                        if nums:
                            val = max(nums) if key in ('rev','cost') else nums[-1]
                            if val > 0:
                                values[key] = round(val)
                                tree_items.append({'label': kw, 'amount': round(val),
                                                   'effective': round(val), 'type': 'item',
                                                   'depth': 0, 'has_children': False,
                                                   'children': []})
                                matched.add(key)
                        break

        if values['cost'] == 0:
            values['cost'] = sum(values[k] for k in ['land','constr','sales','levy','incid','finance'])
        if values['rev'] > 0 and values['cost'] > 0:
            values['profit'] = values['rev'] - values['cost']
            values['roi'] = round(values['profit'] / values['rev'] * 100, 2)

        return {
            'success': True, 'name': name,
            'data': {'name': name, 'tree': tree_items,
                     'line_items': tree_items, 'rev_details': [], 'values': values}
        }
    except Exception as e:
        return {'success': False, 'error': f'PDF 분석 실패: {e}'}


# ────────────────────────────────────────────────────────
# 8. 메인
# ────────────────────────────────────────────────────────
def main():
    # stdout을 UTF-8로 강제 (Windows에서 cp949 출력 방지)
    if hasattr(sys.stdout, 'buffer'):
        sys.stdout = __import__('io').TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

    if len(sys.argv) < 2:
        print(json.dumps({'success': False, 'error': '파일 경로 필요'}, ensure_ascii=False))
        sys.exit(1)

    path = sys.argv[1]
    # sys.argv[2]가 있으면 원본 파일명으로 사용
    # Python 3 on Windows는 GetCommandLineW()로 유니코드 인수를 직접 읽으므로 별도 디코딩 불필요
    orig = sys.argv[2] if len(sys.argv) > 2 else None

    result = (extract_from_pdf(path, orig)
              if path.lower().endswith('.pdf')
              else extract_from_excel(path, orig))
    print(json.dumps(result, ensure_ascii=False))


if __name__ == '__main__':
    main()
